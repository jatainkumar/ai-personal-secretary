"""
Contact Parsers - Extract contacts from various file formats

Handles:
- VCF (vCard) files
- LinkedIn Connections CSV
- Generic CSV files with auto-column detection
- XLSX files (Excel/LinkedIn exports)
"""

import os
import csv
import vobject
import openpyxl
from typing import List, Dict
from langchain_core.documents import Document


def load_vcf_file(file_path: str) -> str:
    """Parse VCF file and return formatted text content"""
    text_content = []
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            for vcard in vobject.readComponents(f):
                fn = vcard.fn.value if hasattr(vcard, 'fn') else 'Unknown'
                
                phones = []
                if hasattr(vcard, 'tel'):
                    for t in vcard.contents.get('tel', []):
                        phones.append(t.value)
                phone_str = ", ".join(phones)

                emails = []
                if hasattr(vcard, 'email'):
                    for e in vcard.contents.get('email', []):
                        emails.append(e.value)
                email_str = ", ".join(emails)
                
                entry = f"Contact Record:\\nName: {fn}\\nPhone: {phone_str}\\nEmail: {email_str}"
                text_content.append(entry)
        return "\\n---\\n".join(text_content)
    except Exception as e:
        print(f"Error parsing VCF {file_path}: {e}")
        return ""


def load_linkedin_connections(file_path: str) -> List[Document]:
    """Parse LinkedIn Connections CSV export"""
    documents = []
    encodings_to_try = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    lines = None
    
    for encoding in encodings_to_try:
        try:
            with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                lines = f.readlines()
            print(f"Successfully read file with encoding: {encoding}")
            break
        except UnicodeDecodeError:
            continue
    
    if lines is None:
        print(f"❌ Failed to read file {file_path} with any encoding")
        return []
            
    # Find the real header row
    start_index = -1
    for i, line in enumerate(lines):
        if "First Name" in line and "Last Name" in line:
            start_index = i
            print(f"Found Header at Line {i+1}: {line.strip()[:50]}...")
            break
    
    if start_index == -1:
        print("Error: Could not find 'First Name' header row.")
        return []

    try:
        csv_data = lines[start_index:]
        reader = csv.DictReader(csv_data)
        
        count = 0
        for row in reader:
            first = row.get("First Name", "").strip()
            last = row.get("Last Name", "").strip()
            
            if not first and not last:
                continue

            company = row.get("Company", "").strip()
            position = row.get("Position", "").strip()
            email = row.get("Email Address", "").strip()
            url = row.get("URL", "").strip()
            connected_on = row.get("Connected On", "").strip()

            content = (
                f"--- LINKEDIN CONTACT CARD ---\\n"
                f"Name: {first} {last}\\n"
                f"Current Role: {position}\\n"
                f"Company: {company}\\n"
                f"Email: {email if email else 'Not listed'}\\n"
                f"LinkedIn Profile: {url}\\n"
                f"Connected Since: {connected_on}\\n"
                f"Source: LinkedIn Connections Export\\n"
            )
            
            metadata = {
                "source": file_path,
                "type": "contact_card",
                "person_name": f"{first} {last}",
                "company": company or "",
                "email": email or ""
            }
            
            documents.append(Document(page_content=content, metadata=metadata))
            count += 1
            
        print(f"Successfully loaded {count} contacts.")
            
    except Exception as e:
        print(f"❌ Critical Error parsing LinkedIn CSV: {e}")
        
    return documents


def extract_contacts_from_xlsx(file_path: str) -> List[Dict]:
    """Extract contact information from XLSX files (LinkedIn exports)"""
    contacts = []
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active
        
        # LinkedIn exports have headers in row 4, scan first 10 rows to find them
        header_row_idx = None
        headers = []
        
        for row_idx in range(1, min(11, sheet.max_row + 1)):
            row_values = [cell.value for cell in sheet[row_idx]]
            row_str = ' '.join([str(v) for v in row_values if v])
            
            if "First Name" in row_str and "Last Name" in row_str:
                header_row_idx = row_idx
                headers = row_values
                print(f"📋 Found XLSX headers in row {row_idx}")
                print(f"📋 Headers: {headers[:7]}")
                break
        
        if not header_row_idx:
            print("❌ No header row found in XLSX")
            wb.close()
            return []
        
        # Map columns
        field_mapping = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            col_lower = str(header).lower().strip()
            
            if 'first' in col_lower and 'name' in col_lower:
                field_mapping['first_name'] = idx
            if 'last' in col_lower and 'name' in col_lower:
                field_mapping['last_name'] = idx
            if 'email' in col_lower or 'e-mail' in col_lower:
                field_mapping['email'] = idx
            if 'phone' in col_lower or 'number' in col_lower:
                field_mapping['phone'] = idx
            if 'company' in col_lower or 'organization' in col_lower:
                field_mapping['company'] = idx
            if 'position' in col_lower or 'title' in col_lower:
                field_mapping['position'] = idx
            if 'url' in col_lower or 'linkedin' in col_lower or 'profile' in col_lower:
                field_mapping['url'] = idx
            if 'connected' in col_lower:
                field_mapping['connected_on'] = idx
        
        print(f"🔍 Field mapping: {field_mapping}")
        
        # Extract contacts from rows after header
        for row in list(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True)):
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue
                
            first_name = row[field_mapping.get('first_name', -1)] if 'first_name' in field_mapping else ''
            last_name = row[field_mapping.get('last_name', -1)] if 'last_name' in field_mapping else ''
            
            if not first_name and not last_name:
                continue
                
            contact = {
                'first_name': str(first_name or '').strip(),
                'last_name': str(last_name or '').strip(),
                'email': str(row[field_mapping.get('email', -1)] or '').strip() if 'email' in field_mapping else '',
                'phone': str(row[field_mapping.get('phone', -1)] or '').strip() if 'phone' in field_mapping else '',
                'company': str(row[field_mapping.get('company', -1)] or '').strip() if 'company' in field_mapping else '',
                'position': str(row[field_mapping.get('position', -1)] or '').strip() if 'position' in field_mapping else '',
                'url': str(row[field_mapping.get('url', -1)] or '').strip() if 'url' in field_mapping else ''
            }
            contacts.append(contact)
        
        wb.close()
        print(f"✅ Extracted {len(contacts)} contacts from XLSX")
        return contacts
        
    except Exception as e:
        print(f"❌ Failed to extract from XLSX: {e}")
        import traceback
        traceback.print_exc()
        return []


def extract_contacts_from_csv(file_path: str) -> List[Dict]:
    """Extract contact information from any CSV file by detecting relevant columns"""
    contacts = []
    encodings_to_try = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    f = None
    
    for encoding in encodings_to_try:
        try:
            f = open(file_path, 'r', encoding=encoding, errors='replace')
            sample = f.read(1024)
            f.seek(0)
            print(f"✅ Successfully opened CSV with encoding: {encoding}")
            break
        except UnicodeDecodeError:
            if f:
                f.close()
            continue
    
    if not f:
        print(f"❌ Failed to open CSV {file_path} with any encoding")
        return []
    
    try:
        sniffer = csv.Sniffer()
        try:
            delimiter = sniffer.sniff(sample, delimiters=',;\\t|').delimiter
            print(f"✅ Detected delimiter: '{delimiter}'")
        except:
            delimiter = ','
            print(f"⚠️ Using fallback delimiter: ','")
        
        # Read all rows first to find the header row
        f.seek(0)
        all_rows = list(csv.reader(f, delimiter=delimiter))
        
        # Find the header row
        header_row = None
        header_row_idx = 0
        expected_fields = ['name', 'first', 'last', 'email', 'company', 'position', 'phone', 'title']
        
        for idx, row in enumerate(all_rows[:10]):
            if not row:
                continue
            row_lower = [cell.lower().strip() for cell in row if cell]
            matches = sum(1 for cell in row_lower if any(field in cell for field in expected_fields))
            if matches >= 2:
                header_row = row
                header_row_idx = idx
                print(f"📋 Found CSV headers in row {idx + 1}: {row}")
                break
        
        if not header_row:
            print(f"⚠️ Using first row as headers")
            header_row = all_rows[0] if all_rows else []
            header_row_idx = 0
            
        # Define field name variations
        name_fields = ['name', 'full name', 'fullname', 'full_name', 'contact name', 'person name', 'contact_name', 'person_name']
        first_name_fields = ['first name', 'firstname', 'first_name', 'fname', 'given name']
        last_name_fields = ['last name', 'lastname', 'last_name', 'lname', 'surname', 'family name']
        email_fields = ['email', 'email address', 'email_address', 'e-mail', 'mail', 'contact email', 'contact_email']
        phone_fields = ['phone', 'phone number', 'phone_number', 'mobile', 'cell', 'telephone', 'contact phone', 'contact_phone', 'tel']
        company_fields = ['company', 'organization', 'organisation', 'org', 'employer', 'firm', 'workplace']
        position_fields = ['position', 'title', 'role', 'job title', 'job_title', 'designation', 'job']
        url_fields = ['url', 'website', 'linkedin', 'profile', 'link', 'web']
        
        # Map column names
        field_mapping = {}
        for idx, col in enumerate(header_row):
            col_lower = col.lower().strip()
            
            if col_lower in first_name_fields:
                field_mapping['first_name'] = idx
            elif col_lower in last_name_fields:
                field_mapping['last_name'] = idx
            elif col_lower in name_fields:
                field_mapping['name'] = idx
            elif col_lower in email_fields:
                field_mapping['email'] = idx
            elif col_lower in phone_fields:
                field_mapping['phone'] = idx
            elif col_lower in company_fields:
                field_mapping['company'] = idx
            elif col_lower in position_fields:
                field_mapping['position'] = idx
            elif col_lower in url_fields:
                field_mapping['url'] = idx
        
        print(f"🔍 Field mapping: {field_mapping}")
        
        # Check if we have name fields
        has_name = ('name' in field_mapping) or ('first_name' in field_mapping or 'last_name' in field_mapping)
        if not has_name:
            print(f"⚠️ No name fields detected in CSV. Skipping contact extraction.")
            f.close()
            return []
        
        # Process rows
        for row in all_rows[header_row_idx + 1:]:
            if not row or len(row) == 0:
                continue
                
            # Extract first and last name
            if 'first_name' in field_mapping or 'last_name' in field_mapping:
                first_name = row[field_mapping['first_name']].strip() if 'first_name' in field_mapping and len(row) > field_mapping['first_name'] else ''
                last_name = row[field_mapping['last_name']].strip() if 'last_name' in field_mapping and len(row) > field_mapping['last_name'] else ''
            elif 'name' in field_mapping:
                name = row[field_mapping['name']].strip() if len(row) > field_mapping['name'] else ''
                if not name:
                    continue
                name_parts = name.split(' ', 1)
                first_name = name_parts[0] if len(name_parts) > 0 else ''
                last_name = name_parts[1] if len(name_parts) > 1 else ''
            else:
                continue
            
            if not first_name and not last_name:
                continue
            
            email = row[field_mapping['email']].strip() if 'email' in field_mapping and len(row) > field_mapping['email'] else ''
            phone = row[field_mapping['phone']].strip() if 'phone' in field_mapping and len(row) > field_mapping['phone'] else ''
            company = row[field_mapping['company']].strip() if 'company' in field_mapping and len(row) > field_mapping['company'] else ''
            position = row[field_mapping['position']].strip() if 'position' in field_mapping and len(row) > field_mapping['position'] else ''
            url = row[field_mapping['url']].strip() if 'url' in field_mapping and len(row) > field_mapping['url'] else ''
            
            contact = {
                'first_name': first_name,
                'last_name': last_name,
                'email': email or None,
                'phone': phone or None,
                'company': company or None,
                'position': position or None,
                'url': url or None,
            }
            
            if any([first_name, last_name, email, phone, company, position]):
                contacts.append(contact)
                
        print(f"✅ Extracted {len(contacts)} contacts from CSV")
        
    except Exception as e:
        print(f"❌ Error extracting contacts from CSV {file_path}: {e}")
        import traceback
        traceback.print_exc()
    finally:
        f.close()
        
    return contacts
